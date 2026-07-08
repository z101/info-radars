from pathlib import Path

import pytest
from openpyxl import load_workbook

from scraper.database import Database
from xlsx_exporter import export_to_xlsx, import_from_xlsx, BASE_COLUMNS, BASE_HEADER_NAMES


@pytest.fixture
def db():
    return Database(":memory:")


def _seed_article(db, topic="Art", author="A. Author"):
    sid = db.create_session()
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc).isoformat()
    db.upsert_article(2026, 4, "Section", topic, author, "5", "http://radio.ru/a/", sid, now)
    return db._fetchone("SELECT id FROM articles WHERE topic=?", (topic,))["id"]


class TestXlsxExport:
    def test_export_creates_file(self, tmp_path, db):
        _seed_article(db)
        articles = db.get_articles_for_export("all")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))
        assert out.exists()

    def test_export_header(self, tmp_path, db):
        _seed_article(db)
        articles = db.get_articles_for_export("all")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))
        wb = load_workbook(str(out))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        expected = [BASE_HEADER_NAMES[c] for c in BASE_COLUMNS]
        assert header == expected

    def test_export_column_order(self, tmp_path, db):
        _seed_article(db)
        articles = db.get_articles_for_export("all")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))
        wb = load_workbook(str(out))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        assert header[:3] == ["id", "I", "R"]
        assert header[3] == "Date"

    def test_export_data_columns(self, tmp_path, db):
        aid = _seed_article(db, topic="My Topic", author="A. Test")
        articles = db.get_articles_for_export("all")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))
        wb = load_workbook(str(out))
        ws = wb.active
        col_map = {cell.value: idx for idx, cell in enumerate(ws[1])}
        date_cell = str(ws.cell(2, col_map["Date"] + 1).value)
        assert '"2026-04"' in date_cell
        assert ws.cell(2, col_map["URL"] + 1).value is not None
        assert ws.cell(2, col_map["Topic"] + 1).value == "My Topic"

    def test_export_marked_interesting(self, tmp_path, db):
        aid = _seed_article(db)
        db.mark_interesting([aid])
        articles = db.get_articles_for_export("all")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))
        wb = load_workbook(str(out))
        ws = wb.active
        i_col = list(BASE_COLUMNS).index("is_interesting") + 1
        assert ws.cell(2, i_col).value == "Y"

    def test_import_updates_flags(self, tmp_path, db):
        aid = _seed_article(db)
        articles = db.get_articles_for_export("all")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))

        wb = load_workbook(str(out))
        ws = wb.active
        i_col = list(BASE_COLUMNS).index("is_interesting") + 1
        r_col = list(BASE_COLUMNS).index("is_read") + 1
        ws.cell(2, i_col).value = 1
        ws.cell(2, r_col).value = 1
        wb.save(str(out))

        result = import_from_xlsx(str(out), db)
        row = db._fetchone("SELECT is_interesting, is_read FROM articles WHERE id=?", (aid,))
        assert row["is_interesting"] == 1
        assert row["is_read"] == 1
        assert result["updated_interesting"] == 1
        assert result["updated_read"] == 1

    def test_import_no_changes(self, tmp_path, db):
        aid = _seed_article(db)
        db.mark_interesting([aid])
        db.mark_read([aid])
        articles = db.get_articles_for_export("all")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))

        result = import_from_xlsx(str(out), db)
        assert result["updated_interesting"] == 0
        assert result["updated_read"] == 0

    def test_import_file_not_found(self):
        with pytest.raises(FileNotFoundError):
            import_from_xlsx("nonexistent.xlsx", None)