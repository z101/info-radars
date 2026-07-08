import json
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

BASE_COLUMNS = [
    "id", "is_interesting", "is_read", "date", "url",
    "pdf_url", "page", "section", "author", "excerpt", "topic",
]

BASE_HEADER_NAMES = {
    "id": "id", "is_interesting": "I", "is_read": "R",
    "date": "Date", "url": "URL", "pdf_url": "PDF",
    "page": "Page", "section": "Section", "author": "Author",
    "excerpt": "Excerpt", "topic": "Topic",
}

BASE_EDITABLE = {"is_interesting", "is_read"}

SEARCH_COLUMNS = [
    "id", "is_interesting", "is_read", "score", "date", "url",
    "pdf_url", "page", "section", "author", "excerpt", "topic",
    "topical_relevance", "technical_depth",
    "practical_applicability", "novelty", "historical_value",
]

SEARCH_HEADER_NAMES = {
    "id": "id", "is_interesting": "I", "is_read": "R",
    "score": "Score", "date": "Date", "url": "URL",
    "pdf_url": "PDF", "page": "Page", "section": "Section",
    "author": "Author", "excerpt": "Excerpt", "topic": "Topic",
    "topical_relevance": "Topical Relevance",
    "technical_depth": "Technical Depth",
    "practical_applicability": "Practical Applicability",
    "novelty": "Novelty",
    "historical_value": "Historical Value",
}

SEARCH_EDITABLE = {"is_interesting", "is_read"}
EDITABLE_HEADERS = {"I", "R"}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
EDITABLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def export_to_xlsx(articles: list[dict], output_path: str, columns=None, header_names=None, editable=None) -> str:
    if columns is None:
        columns = BASE_COLUMNS
    if header_names is None:
        header_names = BASE_HEADER_NAMES
    if editable is None:
        editable = BASE_EDITABLE

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Articles"

    header = [header_names[c] for c in columns]
    for col_idx, col_name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    for row_idx, article in enumerate(articles, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = article.get(col_name, "")
            if col_name in ("is_interesting", "is_read"):
                val = "Y" if val else ""
            elif col_name in ("url", "pdf_url") and val:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = f'=HYPERLINK("{val}", "{val}")'
                cell.border = THIN_BORDER
                continue
            elif col_name == "date":
                month_url = article.get("month_url", "")
                if month_url:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = f'=HYPERLINK("{month_url}", "{val}")'
                    cell.border = THIN_BORDER
                    continue
            elif val is None:
                val = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER

            if col_name in editable:
                cell.fill = EDITABLE_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name in ("excerpt", "topic"):
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )

    col_widths = {
        "id": 5, "is_interesting": 3, "is_read": 3,
        "date": 8, "url": 5, "pdf_url": 5, "page": 5,
        "section": 15, "author": 15,
        "excerpt": 50, "topic": 50,
        "score": 6,
        "topical_relevance": 12, "technical_depth": 12,
        "practical_applicability": 14, "novelty": 8, "historical_value": 10,
    }
    for col_idx, col_name in enumerate(columns, 1):
        w = col_widths.get(col_name, 10)
        ws.column_dimensions[chr(64 + col_idx)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(str(path))
    logger.info("Exported %d articles to %s", len(articles), path)
    return str(path)


def import_from_xlsx(filepath: str, db) -> dict:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    wb = load_workbook(str(path), data_only=True)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    col_map = {name: idx for idx, name in enumerate(header)}

    for col in EDITABLE_HEADERS:
        if col not in col_map:
            raise ValueError(f"Required column '{col}' not found in {filepath}")
    if "id" not in col_map:
        raise ValueError("Required column 'id' not found")

    updates_interesting = []
    updates_read = []

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row in rows:
        article_id = row[col_map["id"]]
        if article_id is None:
            continue

        xlsx_interesting = bool(row[col_map["I"]])
        xlsx_read = bool(row[col_map["R"]])

        current = db._fetchone(
            "SELECT is_interesting, is_read FROM articles WHERE id = ?",
            (article_id,),
        )
        if current is None:
            logger.warning("Article id=%s not found in DB, skipping", article_id)
            continue

        if xlsx_interesting != bool(current["is_interesting"]):
            updates_interesting.append(article_id)
        if xlsx_read != bool(current["is_read"]):
            updates_read.append(article_id)

    if updates_interesting:
        db.mark_interesting(updates_interesting)
    if updates_read:
        db.mark_read(updates_read)

    result = {
        "total_rows": len(rows),
        "updated_interesting": len(updates_interesting),
        "updated_read": len(updates_read),
    }
    logger.info(
        "Import from %s: %d rows, %d is_interesting, %d is_read",
        filepath, result["total_rows"], result["updated_interesting"], result["updated_read"],
    )
    return result