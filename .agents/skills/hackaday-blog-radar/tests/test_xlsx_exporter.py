import json
import sqlite3
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from database import Database
from xlsx_exporter import export_to_xlsx, import_from_xlsx, BASE_COLUMNS, BASE_HEADER_NAMES


@pytest.fixture
def db():
    return Database(":memory:")


def _seed_article(db, url="https://hackaday.com/a/", content="content v1", summary_ru=None):
    from datetime import datetime, timezone
    sid = db.create_session("led-hacks")
    now = datetime.now(timezone.utc).isoformat()
    db.upsert_article("led-hacks", "Art", url, sid, now, "2024-01-01", "exc", ["led"])
    row = db._fetchone("SELECT id FROM articles WHERE url=?", (url,))
    aid = row["id"]
    db.update_article_full_text(aid, "<raw>", content, sid)
    if summary_ru:
        db.save_summary(aid, summary_ru)
    return aid, sid


class TestXlsxExport:
    def test_export_creates_file(self, tmp_path, db):
        _seed_article(db)
        articles = db.get_articles_for_export("led-hacks")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))
        assert out.exists()

    def test_export_header(self, tmp_path, db):
        _seed_article(db)
        articles = db.get_articles_for_export("led-hacks")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))
        wb = load_workbook(str(out))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        expected = [BASE_HEADER_NAMES[c] for c in BASE_COLUMNS]
        assert header == expected

    def test_export_data(self, tmp_path, db):
        aid, _ = _seed_article(db)
        articles = db.get_articles_for_export("led-hacks")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))
        wb = load_workbook(str(out))
        ws = wb.active
        row2 = [cell.value for cell in ws[2]]
        assert row2[0] == aid
        assert row2[1] is None or row2[1] == ""
        assert row2[2] is None or row2[2] == ""

    def test_export_marked_interesting(self, tmp_path, db):
        aid, _ = _seed_article(db)
        db.mark_interesting([aid])
        articles = db.get_articles_for_export("led-hacks")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))
        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.cell(2, 2).value == "Y"

    def test_export_hyperlink(self, tmp_path, db):
        _seed_article(db)
        articles = db.get_articles_for_export("led-hacks")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))
        wb = load_workbook(str(out))
        ws = wb.active
        url_col = list(BASE_COLUMNS).index("url") + 1
        val = ws.cell(2, url_col).value
        assert val is not None
        assert "HYPERLINK" in str(val)
        assert "hackaday.com" in str(val)
        # Link text should be the URL itself, not a static string
        assert "https://hackaday.com/a/" in str(val)

    def test_export_summary_word_wrap(self, tmp_path, db):
        long_summary = "This is a very long Russian summary that should wrap " * 5
        _seed_article(db, summary_ru=long_summary)
        articles = db.get_articles_for_export("led-hacks")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))
        wb = load_workbook(str(out))
        ws = wb.active
        summary_col = list(BASE_COLUMNS).index("summary_ru") + 1
        cell = ws.cell(2, summary_col)
        assert cell.alignment.wrap_text is True
        assert cell.value == long_summary

    def test_import_updates_flags(self, tmp_path, db):
        aid, _ = _seed_article(db, url="https://hackaday.com/b/")
        articles = db.get_articles_for_export("led-hacks")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))

        wb = load_workbook(str(out))
        ws = wb.active
        ws.cell(2, 2).value = 1
        ws.cell(2, 3).value = 1
        wb.save(str(out))

        result = import_from_xlsx(str(out), db)
        row = db._fetchone("SELECT is_interesting, is_read FROM articles WHERE id=?", (aid,))
        assert row["is_interesting"] == 1
        assert row["is_read"] == 1
        assert result["updated_interesting"] == 1
        assert result["updated_read"] == 1

    def test_import_no_changes(self, tmp_path, db):
        aid, _ = _seed_article(db, url="https://hackaday.com/c/")
        db.mark_interesting([aid])
        db.mark_read([aid])
        articles = db.get_articles_for_export("led-hacks")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))

        result = import_from_xlsx(str(out), db)
        assert result["updated_interesting"] == 0
        assert result["updated_read"] == 0

    def test_import_file_not_found(self):
        with pytest.raises(FileNotFoundError):
            import_from_xlsx("nonexistent.xlsx", None)

    def test_tags_flattened(self, tmp_path, db):
        from datetime import datetime, timezone
        sid = db.create_session("led-hacks")
        now = datetime.now(timezone.utc).isoformat()
        db.upsert_article("led-hacks", "Tagged Art", "https://tagged.url/", sid, now,
                          "2024-06-01", "exc", ["led", "esp32", "matrix"])
        row = db._fetchone("SELECT id FROM articles WHERE url='https://tagged.url/'")
        articles = db.get_articles_for_export("led-hacks")
        out = tmp_path / "test.xlsx"
        export_to_xlsx(articles, str(out))
        wb = load_workbook(str(out))
        ws = wb.active
        tags_col = list(BASE_COLUMNS).index("tags") + 1
        val = ws.cell(ws.max_row, tags_col).value
        assert val == "led, esp32, matrix"